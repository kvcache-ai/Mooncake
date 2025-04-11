import sys
import json
import os
import openpyxl

global metrics
metrics = ['request_throughput', 'output_throughput', 'total_token_throughput',\
            'mean_ttft_ms', 'median_ttft_ms', 'std_ttft_ms', 'p99_ttft_ms',\
            'mean_tpot_ms', 'median_tpot_ms', 'std_tpot_ms', 'p99_tpot_ms',\
            'mean_itl_ms', 'median_itl_ms', 'std_itl_ms', 'p99_itl_ms', \
            'mean_e2el_ms', 'median_e2el_ms', 'std_e2el_ms', 'p99_e2el_ms' ]

def parse_serving_throughput(path: str):
    values=[]
    with open(path, 'r') as f:
        result = json.load(f)
        for metric in metrics:
            value = result[metric]
            values.append(value)
    return values
            

if __name__ == '__main__':
    if len(sys.argv) != 3:
        print("Usage: {} <result_path> <parsed_result_path>".format(sys.argv[0]))
        sys.exit(1)
    result_path = sys.argv[1]
    parsed_result_path = sys.argv[2]
    col=1
    row=1
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    config_names=['num_pserver','num_dserver','input_len','output_len','max_concurrency']
    for con in config_names:
        sheet.cell(row,col,con)
        row += 1
    for metric in metrics:
        sheet.cell(row,col,metric)
        row += 1
    files = os.listdir(result_path)
    files.sort()
    for file in files:
        if file.endswith("json"):
            configs=file.split('-')
            col += 1
            sheet.cell(1,col,configs[1])
            sheet.cell(2,col,configs[2])
            sheet.cell(3,col,configs[4])
            sheet.cell(4,col,configs[6])
            sheet.cell(5,col,configs[8])
            results=parse_serving_throughput(os.path.join(result_path,file))
            row=5
            for result in results:
                row+=1
                sheet.cell(row,col,result)
    workbook.save(parsed_result_path)